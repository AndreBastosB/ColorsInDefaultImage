from collections import Counter
from sklearn.cluster import KMeans
from matplotlib import colors
import matplotlib.pyplot as plt
import numpy as np
import cv2
import os 
import xlwt
from tempfile import TemporaryFile
import pandas as pd
from openpyxl import load_workbook
import time
import pyexcel as p
from PIL import Image
import shutil
import pytesseract

directory1 = (r'C:\Users\Take4\Desktop\Raias')

#lista de arquivos nas pastas
list1 = os.listdir(directory1)

#remove ext from files to list them
myList1 = [i.split('.png')[0] for i in list1]

#funcao para enviar arquvios que nao existem nas duas pastas para um CSV
def exportCSV(wishList):
    book = xlwt.Workbook()
    #sheet name below
    sheet1 = book.add_sheet('sheet1')

    for i,e in enumerate(wishList):
        sheet1.write(i,1,e)

    #xls file name below
    name = "Aaa.xls"
    book.save(name)
    book.save(TemporaryFile())

exportCSV(myList1)

#------------------------------------------------------------------------------
#CONTAGEM DA QUANTIDADE DE ARQUIVOS QUE SERÃO EXAMINADOS NA PASTA
#------------------------------------------------------------------------------

path, dirs, files = next(os.walk(directory1))
file_count = len(files)

#------------------------------------------------------------------------------
#TRANSFORMAR XLS EM XLSX
#------------------------------------------------------------------------------

p.save_book_as (file_name= r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xls',
               dest_file_name= r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xlsx')

os.remove(r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xls')

#------------------------------------------------------------------------------

linha = 1
coluna = 2
path = r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xlsx'

while linha <= file_count:
    
    path2 = load_workbook(path)
    path3 = path2 ['sheet1']
    valor = path3.cell(row=linha, column=coluna).value

    im = Image.open(r'C:\Users\Take4\Desktop\Raias\\'+valor+'.png').convert("RGB")
    
    # get pixels
    pixels = [im.getpixel((i, j)) for j in range(im.height) for i in range(im.width)]
    
    # or
    pixels = [i for i in im.getdata()]
    
    #check if tuple of pixel value exists in array-pixel
    
    
    check6 = ((254, 254, 254) in pixels) 
    check7 = ((220, 220, 220) in pixels)
    check2 = ((208, 208, 208) in pixels)
    check3 = ((210, 0, 1) in pixels)
    
    
    check1 = ((181, 1, 2) in pixels)
    check4 = ((244, 0, 0) in pixels)
    
    
    if check6 == False or check7 == False or check2 == False or check3 == False:
        print (valor)
    else: 
        print ('\n')
    
    # print (valor)
    # print (check1)
    # print (check4)

    # print ('\n')
    
    
    linha += 1

