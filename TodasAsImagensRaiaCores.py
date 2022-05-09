from collections import Counter
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt
import cv2
import os 
import xlwt
from tempfile import TemporaryFile
from openpyxl import load_workbook
import pyexcel as p

directory1 = (r'C:\Users\Take4\Desktop\Raias')

#lista de arquivos nas pastas
list1 = os.listdir(directory1)

#remove ext from files to list them
myList1 = [i.split('.png')[0] for i in list1]

#funcao para enviar arquvios que nao existem nas duas pastas para um CSV
def exportCSV(wishList):
    book = xlwt.Workbook()
    #sheet name below
    sheet1 = book.add_sheet('sheet1')

    for i,e in enumerate(wishList):
        sheet1.write(i,1,e)

    #xls file name below
    name = "Aaa.xls"
    book.save(name)
    book.save(TemporaryFile())

exportCSV(myList1)

#------------------------------------------------------------------------------
#CONTAGEM DA QUANTIDADE DE ARQUIVOS QUE SERÃO EXAMINADOS NA PASTA
#------------------------------------------------------------------------------

path, dirs, files = next(os.walk(directory1))
file_count = len(files)

#------------------------------------------------------------------------------
#TRANSFORMAR XLS EM XLSX
#------------------------------------------------------------------------------

p.save_book_as (file_name= r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xls',
               dest_file_name= r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xlsx')

os.remove(r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xls')

#------------------------------------------------------------------------------

def preprocess(raw):
    image = cv2.resize(raw, (900, 600), interpolation = cv2.INTER_AREA)                                          
    image = image.reshape(image.shape[0]*image.shape[1], 3)
    return image

def rgb_to_hex(rgb_color):
    hex_color = "#"
    for i in rgb_color:
        hex_color += ("{:02x}".format(int(i)))
    return hex_color

lista = []

def analyze(img):
    clf = KMeans(n_clusters = 100)
    color_labels = clf.fit_predict(img)
    center_colors = clf.cluster_centers_
    counts = Counter(color_labels)
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [rgb_to_hex(ordered_colors[i]) for i in counts.keys()]

    plt.figure(figsize = (12, 8))
    plt.pie(counts.values(), labels = hex_colors, colors = hex_colors)

    print("Found the following colors:\n")
    for color in hex_colors:
      print (color)  
      lista.append(color)

linha = 1
coluna = 2
path = r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xlsx'

while linha <= file_count:
    
    path2 = load_workbook(path)
    path3 = path2 ['sheet1']
    valor = path3.cell(row=linha, column=coluna).value

    image = cv2.imread(r'C:\Users\Take4\Desktop\Raias\\' + str(valor + '.png'))
    image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

    modified_image = preprocess(image)
    analyze(modified_image)
    print (linha)
    linha += 1
    
listaFinal = (Counter(lista))

print (listaFinal)

