import os 
import xlwt
from tempfile import TemporaryFile
from openpyxl import load_workbook
import pyexcel as p
from PIL import Image
import pytesseract

#------------------------------------------------------------------------------
# TRANSFORMANDO TODOS OS NOMES DE ARQUIVOS EM EXCEL XLS.
#------------------------------------------------------------------------------

directory1 = (r'C:\Users\Take4\Desktop\3011')
directory2 = (r'C:\Users\Take4\Desktop\Raias')

#lista de arquivos nas pastas
list1 = os.listdir(directory1)

#remove ext from files to list them
myList1 = [i.split('.png')[0] for i in list1]

#funcao para enviar arquvios que nao existem nas duas pastas para um CSV
def exportCSV(wishList):
    book = xlwt.Workbook()
    #sheet name below
    sheet1 = book.add_sheet('sheet1')

    for i,e in enumerate(wishList):
        sheet1.write(i,1,e)

    #xls file name below
    name = "Aaa.xls"
    book.save(name)
    book.save(TemporaryFile())

exportCSV(myList1)

#------------------------------------------------------------------------------
#CONTAGEM DA QUANTIDADE DE ARQUIVOS QUE SERÃO EXAMINADOS NA PASTA
#------------------------------------------------------------------------------

path, dirs, files = next(os.walk(directory1))
file_count = len(files)

#------------------------------------------------------------------------------
#TRANSFORMAR XLS EM XLSX
#------------------------------------------------------------------------------

p.save_book_as (file_name= r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xls',
               dest_file_name= r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xlsx')

os.remove(r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xls')

#------------------------------------------------------------------------------
#HORA DA MAGIA - BASEADO NO EXCEL, TODOS OS ARQUIVOS SÃO VERIFICADOS.
#SE POSSUIREM TODAS AS IMAGENS DA RAIA, A IMAGEM É TRANSFERIDA PARA OUTRA PASTA, POSTERIORMENTE VERIFICADA.
#------------------------------------------------------------------------------

linha = 1
coluna = 2

path = r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xlsx'
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract'

while linha <= file_count:
    
    path2 = load_workbook(path)
    path3 = path2 ['sheet1']
    valor = path3.cell(row=linha, column=coluna).value

    #im = Image.open(r'C:\Users\Take4\Desktop\7898040320652.png').convert("RGB")
    im = Image.open(directory1 + '\\' + str(valor) + '.png').convert("RGB") 

    # get pixels
    pixels = [im.getpixel((i, j)) for j in range(im.height) for i in range(im.width)]
    
    # or
    pixels = [i for i in im.getdata()]
    
    #check if tuple of pixel value exists in array-pixel
    
    check6 = ((254, 254, 254) in pixels) 
    check7 = ((220, 220, 220) in pixels)
    check2 = ((208, 208, 208) in pixels)
    check3 = ((210, 0, 1) in pixels)
    
    #As imagens variam. Ou a imagem tem o check1, ou tem o check4.
    check1 = ((181, 1, 2) in pixels)
    check4 = ((244, 0, 0) in pixels)
    
    #Check de Foto indisponivel
    check6Tessreact = (pytesseract.image_to_string(directory1 + '\\' + str(valor) + '.png')) #foto indisponivel
    
    if check6 == True and check7 == True and check2 == True and check3 == True and (check1 == True or check4 == True):
        os.replace(directory1 + '\\' + str(valor) + '.png', directory2 + '\\' + str(valor) + '.png')
    elif 'foto indisponivel' in check6Tessreact:
        os.replace(directory1 + '\\' + str(valor) + '.png', directory2 + '\\' + str(valor) + '.png')
    else:
        pass

    
    print (valor)
    print (linha)
    linha +=1 
    
os.remove(r'C:\Users\Take4\Desktop\PythonCodes\Checking S3 Crawler\Aaa.xlsx')
    
    

















